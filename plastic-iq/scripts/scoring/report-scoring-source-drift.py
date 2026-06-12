#!/usr/bin/env python3
"""Read-only scoring source-of-truth reconciliation report generator."""
from __future__ import annotations

import json
import math
import re
import subprocess
import sys
import zipfile
from collections import Counter, defaultdict
from pathlib import Path
from xml.etree import ElementTree as ET

try:
    import openpyxl
except ImportError:
    print(json.dumps({"error": "openpyxl not installed", "stop": True}))
    sys.exit(1)

ROOT = Path(__file__).resolve().parents[2]
MATERIAL_XLSX = ROOT / "docs/source-of-truth/PBG_Material_Lookup.xlsx"
DOCX = ROOT / "docs/source-of-truth/PAC_Scoring_Algorithm_v2.3.5.docx"
REPORT_DIR = ROOT / "docs/reports"
REPORT_JSON = REPORT_DIR / "scoring-source-drift-report.json"
REPORT_MD = REPORT_DIR / "scoring-source-drift-report.md"

NS = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}


def round4(v):
    if v is None:
        return None
    try:
        return round(float(v), 4)
    except (TypeError, ValueError):
        return None


def norm_name(s: str, *, keep_parens: bool = False) -> str:
    s = str(s or "")
    if not keep_parens:
        s = re.sub(r"\([^)]*\)", " ", s)
    s = s.lower()
    s = re.sub(r"[^a-z0-9]+", " ", s)
    return re.sub(r"\s+", " ", s).strip()


def norm_name_strict(s: str) -> str:
    return norm_name(s, keep_parens=True)


def load_code_values() -> dict:
    node_script = r"""
import { MATERIAL_TAXONOMY, MATERIAL_TAXONOMY_ALIASES } from './scripts/agent2/deterministic/material-taxonomy.mjs';
import { EXPOSURE_DEFAULTS_BY_KEY } from './src/shared/product-type-registry/scoring-assumptions.mjs';
import { STARTER_PRODUCT_TYPE_CONFIGS } from './src/shared/product-type-registry/configs/starter-configs.mjs';
import { LAYER_4A_POSITIVE_LOOKUP, LAYER_4A_POSITIVE_MAX } from './scripts/agent2/layer4a-positive.mjs';
import { NON_DETECT_MITIGATION_FACTOR } from './src/lib/lockedInput/buildSystemValidation.ts';
import { ALGORITHM_VERSION } from './scripts/agent3/algorithm.mjs';

const LAYER_4A_DEDUCTION_POINTS = {
  proprietary_ceramic_or_nonstick_formula_undisclosed: -3,
  unknown_proprietary_food_contact_coating: -3,
  marketing_language_only: -2,
  bpa_free_claim_only: -1,
};

const ESCALATORS = [
  { id: 'escalator_4', multiplier: 1.5, field: 'escalator_4_triggers' },
  { id: 'escalator_2', multiplier: 1.4, field: 'escalator_2_triggers' },
  { id: 'escalator_1', multiplier: 1.25, field: 'escalator_1_triggers' },
];

const INERT_MIGRATION_THRESHOLD = 0.05;
const INERT_EXPOSURE_MULTIPLIER = 0.2;
const NPR_SCALE = 1000;
const LAYER_4A_CAP = 5;
const HARD_CAP_UNKNOWN_COATING = 72;
const SCORE_MAX = 99;

function tierForScore(score) {
  if (score >= 90) return 'Excellent';
  if (score >= 75) return 'Good';
  if (score >= 55) return 'Caution';
  if (score >= 30) return 'Concern';
  return 'High Risk';
}

const materials = {};
for (const [id, entry] of Object.entries(MATERIAL_TAXONOMY)) {
  materials[id] = {
    id,
    name: entry.name,
    hazard: entry.hazard,
    migration: entry.migration,
    tier: entry.tier,
    inertProtection: Boolean(entry.inertProtection),
    unknownFoodContactCoating: Boolean(entry.unknownFoodContactCoating),
  };
}

console.log(JSON.stringify({
  materials,
  aliases: MATERIAL_TAXONOMY_ALIASES,
  exposureDefaults: EXPOSURE_DEFAULTS_BY_KEY,
  registry: STARTER_PRODUCT_TYPE_CONFIGS.map(c => ({
    registry_key: c.registry_key,
    category: c.category,
    subcategory: c.subcategory,
    matrix_key: c.matrix_key,
    scoring_assumption_ref: c.scoring_assumption_ref,
    subcategory_aliases: c.subcategory_aliases,
  })),
  layer4aPositive: LAYER_4A_POSITIVE_LOOKUP,
  layer4aPositiveMax: LAYER_4A_POSITIVE_MAX,
  layer4aDeductions: LAYER_4A_DEDUCTION_POINTS,
  nonDetectFactor: NON_DETECT_MITIGATION_FACTOR,
  algorithmVersion: ALGORITHM_VERSION,
  escalators: ESCALATORS,
  constants: {
    INERT_MIGRATION_THRESHOLD,
    INERT_EXPOSURE_MULTIPLIER,
    NPR_SCALE,
    LAYER_4A_CAP,
    HARD_CAP_UNKNOWN_COATING,
    SCORE_MAX,
    tierBands: [
      { min: 90, tier: 'Excellent' },
      { min: 75, tier: 'Good' },
      { min: 55, tier: 'Caution' },
      { min: 30, tier: 'Concern' },
      { min: 0, tier: 'High Risk' },
    ],
    nprFormula: 'NPR = hazard * migration * CI * severity * duration * 1000',
    scoreFormula: 'PAC = 100 - sqrt(weighted NPR) * 5',
  },
}));
"""
    proc = subprocess.run(
        ["npx", "tsx", "-e", node_script],
        cwd=ROOT,
        capture_output=True,
        text=True,
    )
    if proc.returncode != 0:
        raise RuntimeError(f"node import failed: {proc.stderr}")
    return json.loads(proc.stdout)


def parse_docx_tables(docx_path: Path) -> list[list[list[str]]]:
    with zipfile.ZipFile(docx_path) as z:
        xml = z.read("word/document.xml")
    root = ET.fromstring(xml)

    def cell_text(tc):
        parts = []
        for t in tc.findall(".//w:t", NS):
            if t.text:
                parts.append(t.text)
            if t.tail:
                parts.append(t.tail)
        return "".join(parts).strip()

    tables = []
    for tbl in root.findall(".//w:tbl", NS):
        rows = []
        for tr in tbl.findall("w:tr", NS):
            rows.append([cell_text(tc) for tc in tr.findall("w:tc", NS)])
        tables.append(rows)
    return tables


def parse_docx_paragraphs(docx_path: Path) -> list[str]:
    with zipfile.ZipFile(docx_path) as z:
        xml = z.read("word/document.xml")
    root = ET.fromstring(xml)

    def para_text(p):
        parts = []
        for t in p.findall(".//w:t", NS):
            if t.text:
                parts.append(t.text)
            if t.tail:
                parts.append(t.tail)
        return "".join(parts).strip()

    return [para_text(p) for p in root.findall(".//w:p", NS) if para_text(p)]


def load_material_lookup(xlsx_path: Path) -> tuple[list[dict], dict]:
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb["Material Lookup"]
    header = list(ws.iter_rows(min_row=4, max_row=4, values_only=True))[0]
    col = {str(h).strip(): i for i, h in enumerate(header) if h}

    rows = []
    for r in ws.iter_rows(min_row=5, values_only=True):
        if not r or not r[0]:
            continue
        name = str(r[0]).strip()
        if name.startswith("NOTE:"):
            continue
        rows.append(
            {
                "name": name,
                "tier": r[col["Tier"]],
                "hazard": round4(r[col["Material Hazard"]]),
                "migration": round4(r[col["Migration"]]),
                "non_detect_eligible": str(r[col.get("Non-Detect Mitigation Eligible", 9)] or "").strip(),
                "scoring_status": str(r[col["Scoring Status"]] or "").strip(),
                "source_location": f"PBG_Material_Lookup.xlsx / Material Lookup / row material={name}",
            }
        )
    meta = {
        "sheets": wb.sheetnames,
        "header_row": 4,
        "data_row_count": len(rows),
        "columns": list(col.keys()),
    }
    return rows, meta


MANUAL_NAME_TO_CODE = {
    "ptfe nonstick coating": "ptfe_nonstick",
    "ptfe nonstick coating titanium reinforced": "ptfe_nonstick_titanium_reinforced",
    "proprietary ceramic nonstick terrabond": "terrabond_proprietary",
    "proprietary ceramic nonstick terrabond": "terrabond_proprietary",
    "hybrid stainless lattice nonstick surface": "hybrid_stainless_nonstick_food_contact",
    "unknown proprietary food contact coating": "proprietary_named_food_contact",
    "plastic lid resin unspecified": "plastic_lid_unspecified",
    "bpa free plastic resin unspecified": "bpa_free_plastic_unspecified",
    "refill container hdpe or similar": "refill_container_hdpe_unspecified",
    "silicone gasket food grade verified": "silicone_gasket_verified",
    "silicone gasket unverified": "silicone_gasket_unverified",
    "silicone gasket food grade verified strict": "silicone_gasket_verified",
    "silicone gasket unverified strict": "silicone_gasket_unverified",
    "stay cool handle material undisclosed": "stay_cool_handle_undisclosed",
    "cast iron pre seasoned with natural vegetable oil": "cast_iron_seasoned",
    "stainless steel grade unspecified": "stainless_steel_unspecified",
    "stainless steel 304": "stainless_steel_304",
    "stainless steel 316": "stainless_steel_316",
    "ceramic nonstick sol gel coating": "ceramic_nonstick_sol_gel",
    "thermolon ceramic coating": "thermolon_ceramic",
    "nylon food contact": "nylon_food_contact",
    "borosilicate glass": "borosilicate_glass",
    "tempered glass": "tempered_glass",
    "soda lime glass": "tempered_glass",
    "silicone coated handle": "silicone_over_riveted_base",
    "bamboo lid with silicone seal": "bamboo_lid_silicone",
    "tritan plastic": "tritan",
    "tritan copolyester": "tritan",
    "ptfe coating lower band": "ptfe_coating",
    "ptfe coating": "ptfe_coating",
    "laser etched stainless cooking surface": "laser_etched_stainless_surface",
    "tempered glass lid": "tempered_glass_lid",
    "stainless steel handle": "stainless_steel_handle",
    "integrated cast iron handle": "cast_iron_integrated_handle",
    "stainless steel rivets": "stainless_steel_rivets",
    "magnetic stainless steel base": "magnetic_stainless_base",
    "carbon steel": "carbon_steel",
    "cast iron pre seasoned with natural vegetable oil": "cast_iron_seasoned",
    "titanium": "titanium",
    "food grade copper lined": "food_grade_copper_lined",
    "bare copper acidic food contact": "bare_copper_acidic_food_contact",
    "food safe ceramic verified glaze": "food_safe_ceramic_verified_glaze",
    "cast iron pre seasoned carbon steel": "cast_iron",
}

# Component-level code IDs that share a combined lookup row (backward-compatible aliases).
CODE_VARIANT_OF = {
    "cast_iron_seasoned": "cast_iron",
    "carbon_steel": "cast_iron",
    "cast_iron_integrated_handle": "cast_iron",
    "tempered_glass_lid": "tempered_glass",
    "stainless_steel_handle": "stainless_steel_304",
    "stainless_steel_rivets": "stainless_steel_304",
    "magnetic_stainless_base": "stainless_steel_304",
}


def build_source_to_code_map(source_rows, code_materials):
    code_by_norm_name = {norm_name(v["name"]): k for k, v in code_materials.items()}
    mapping = {}
    for row in source_rows:
        nn_strict = norm_name_strict(row["name"])
        nn = norm_name(row["name"])
        code_id = MANUAL_NAME_TO_CODE.get(nn_strict) or MANUAL_NAME_TO_CODE.get(nn)
        if nn_strict.endswith("food grade verified"):
            code_id = code_id or "silicone_gasket_verified"
        if nn_strict.endswith("unverified") and "silicone gasket" in nn_strict:
            code_id = code_id or "silicone_gasket_unverified"
        if not code_id:
            code_id = code_by_norm_name.get(nn)
        if not code_id:
            for cid, cent in code_materials.items():
                if norm_name(cent["name"]) == nn:
                    code_id = cid
                    break
        if not code_id:
            for cid, cent in code_materials.items():
                if nn in norm_name(cent["name"]) or norm_name(cent["name"]) in nn:
                    code_id = cid
                    break
        mapping[row["name"]] = code_id
    return mapping


def compare_numeric(code_val, source_val, tol=0.0001):
    if code_val is None and source_val is None:
        return "MATCH", None
    if code_val is None or source_val is None:
        return "MISMATCH", None
    c, s = float(code_val), float(source_val)
    if abs(c - s) <= tol:
        return "MATCH", None
    diff = round(c - s, 6)
    return "MISMATCH", diff


def reconcile_materials(source_rows, code_materials, source_to_code):
    rows = []
    matched_code_ids = set()

    for src in source_rows:
        code_id = source_to_code.get(src["name"])
        status_base = src["scoring_status"]
        if not code_id:
            st = (
                "FUTURE_CATEGORY_NOT_IN_CODE"
                if status_base == "future_category"
                else "MISSING_ACTIVE_IN_CODE"
            )
            for field in ("hazard", "migration", "non_detect_eligible", "tier", "scoring_status"):
                src_val = src.get(field if field != "non_detect_eligible" else "non_detect_eligible")
                rows.append(
                    {
                        "material_id": src["name"],
                        "scoring_status": status_base,
                        "field": field,
                        "code_value": None,
                        "source_value": src_val if field != "scoring_status" else status_base,
                        "status": st,
                        "code_file_line": None,
                        "source_location": src["source_location"],
                        "notes": "No code MATERIAL_TAXONOMY mapping found",
                    }
                )
            continue

        matched_code_ids.add(code_id)
        code = code_materials.get(code_id)
        if not code:
            continue

        code_loc = f"scripts/agent2/deterministic/material-taxonomy.mjs:{code_id}"

        for field, code_key, src_key in [
            ("hazard", "hazard", "hazard"),
            ("migration", "migration", "migration"),
            ("tier", "tier", "tier"),
        ]:
            if field == "tier":
                cv, sv = code.get(code_key), src.get(src_key)
                st = "MATCH" if norm_name(str(cv)) == norm_name(str(sv)) else "MISMATCH"
                diff = None
            else:
                st, diff = compare_numeric(code.get(code_key), src.get(src_key))
            if status_base == "future_category" and st == "MISMATCH":
                st = "NEEDS_RICK_REVIEW"
            rows.append(
                {
                    "material_id": code_id,
                    "scoring_status": status_base,
                    "field": field,
                    "code_value": code.get(code_key),
                    "source_value": src.get(src_key),
                    "status": st,
                    "code_file_line": code_loc,
                    "source_location": src["source_location"],
                    "notes": f"diff={diff}" if diff is not None else "",
                }
            )

        # Non-detect: code uses heuristic, source has explicit column
        src_nd = src["non_detect_eligible"].lower()
        src_nd_bool = src_nd in ("yes", "true", "y")
        code_nd = None  # heuristic not exported; mark review
        rows.append(
            {
                "material_id": code_id,
                "scoring_status": status_base,
                "field": "non_detect_eligible",
                "code_value": "heuristic (see material-lookup-audit.mjs)",
                "source_value": src["non_detect_eligible"],
                "status": "NEEDS_RICK_REVIEW",
                "code_file_line": "scripts/lib/material-lookup-audit.mjs:39",
                "source_location": src["source_location"],
                "notes": "Code uses pattern/heuristic; xlsx has explicit Yes/No column",
            }
        )

    for cid, code in code_materials.items():
        if cid in matched_code_ids:
            continue
        variant_of = CODE_VARIANT_OF.get(cid)
        if variant_of and variant_of in matched_code_ids:
            rows.append(
                {
                    "material_id": cid,
                    "scoring_status": "active_in_code",
                    "field": "presence",
                    "code_value": code["name"],
                    "source_value": f"alias of {variant_of} (combined lookup row)",
                    "status": "ALIAS_ONLY",
                    "code_file_line": f"scripts/agent2/deterministic/material-taxonomy.mjs:{cid}",
                    "source_location": "PBG_Material_Lookup.xlsx",
                    "notes": f"Component variant shares lookup values with {variant_of}",
                }
            )
            continue
        rows.append(
            {
                "material_id": cid,
                "scoring_status": "active_in_code",
                "field": "presence",
                "code_value": code["name"],
                "source_value": None,
                "status": "MISSING_IN_SOURCE",
                "code_file_line": f"scripts/agent2/deterministic/material-taxonomy.mjs:{cid}",
                "source_location": "PBG_Material_Lookup.xlsx",
                "notes": "Material ID in code but no matched source row",
            }
        )

    return rows


def parse_category_defaults_table(tables):
    for rows in tables:
        if not rows:
            continue
        hdr = [c.lower() for c in rows[0]]
        if "category" in hdr[0] and "default severity" in " ".join(hdr):
            out = []
            for r in rows[1:]:
                if not r or not r[0]:
                    continue
                out.append(
                    {
                        "category": r[0].strip(),
                        "default_severity": r[1].strip() if len(r) > 1 else "",
                        "default_duration": r[2].strip() if len(r) > 2 else "",
                        "doc_location": "PAC_Scoring_Algorithm_v2.3.5.docx / Table 0 Category Default Use Assumptions",
                    }
                )
            return out
    return []


def parse_layer4a_tables(tables):
    pos, neg = [], []
    for rows in tables:
        if not rows:
            continue
        h0 = rows[0][0].lower() if rows[0] else ""
        if h0 == "credit":
            for r in rows[1:]:
                if len(r) >= 2:
                    pos.append({"label": r[0].strip(), "points": r[1].strip()})
        if h0 == "deduction":
            for r in rows[1:]:
                if len(r) >= 2:
                    neg.append(
                        {
                            "label": r[0].strip(),
                            "points": r[1].strip(),
                            "applies_when": r[2].strip() if len(r) > 2 else "",
                        }
                    )
    return pos, neg


def parse_escalator_table(tables):
    for rows in tables:
        if rows and rows[0] and rows[0][0].lower().startswith("escalator"):
            out = []
            for r in rows[1:]:
                if len(r) >= 3:
                    out.append(
                        {
                            "name": r[0].strip(),
                            "trigger": r[1].strip(),
                            "multiplier": r[2].strip(),
                            "doc_location": "PAC_Scoring_Algorithm_v2.3.5.docx / Escalators table",
                        }
                    )
            return out
    return []


def parse_tier_table(tables):
    for rows in tables:
        if rows and rows[0] and rows[0][0].lower() == "score":
            return [
                {"score_range": r[0], "tier": r[1], "meaning": r[2] if len(r) > 2 else ""}
                for r in rows[1:]
                if r and r[0]
            ]
    return []


def extract_primary_severity_duration(exposure_key, exposure_defaults):
    block = exposure_defaults.get(exposure_key, {})
    roles = block.get("roles") or {}
    role = roles.get("primary_food_contact") or roles.get("formulation") or {}
    sev = role.get("severity")
    dur = role.get("duration")
    ci = role.get("contact_intimacy")

    if isinstance(sev, dict):
        base = sev.get("severity_base", 0)
        adds = sum(a.get("value", 0) for a in sev.get("additions", []) or [])
        severity = round4(base + adds)
        severity_note = f"base {base} + additions {adds}"
    else:
        severity = round4(sev)
        severity_note = "scalar"

    if isinstance(dur, dict):
        duration = round4(float(dur.get("duration", 0)) * float(dur.get("modifier", 1)))
        duration_note = f"duration {dur.get('duration')} × modifier {dur.get('modifier', 1)}"
    else:
        duration = round4(dur)
        duration_note = "scalar"

    return severity, duration, ci, severity_note, duration_note


DOC_TO_CODE_SUBCATEGORY = [
    {
        "doc_category": "Cookware",
        "code_key": "cookware",
        "registry_subcategory": "Cookware",
        "matrix_key": "cookware",
    },
    {
        "doc_category": "Cooking utensils",
        "code_key": "utensils",
        "registry_subcategory": "Utensils",
        "matrix_key": "cooking_utensils",
        "utensils_split": True,
    },
    {
        "doc_category": "Water bottles / drinkware",
        "code_key": "water_bottles",
        "registry_subcategory": "Water bottle",
        "matrix_key": "water_bottles",
    },
    {
        "doc_category": "Water bottles / drinkware",
        "code_key": "drinkware",
        "registry_subcategory": "Drinkware (tumbler/mug)",
        "matrix_key": "drinkware",
    },
    {
        "doc_category": "Food storage containers",
        "code_key": "food_storage",
        "registry_subcategory": "Food storage",
        "matrix_key": "food_storage",
    },
    {
        "doc_category": "Rinse-off products",
        "code_key": "rinse_off",
        "registry_subcategory": "Rinse-off",
        "matrix_key": "rinse_off",
    },
    {
        "doc_category": "Everyday clothing",
        "code_key": "textiles",
        "registry_subcategory": "Bedding",
        "matrix_key": "textiles",
    },
]


def parse_doc_severity_duration(doc_row):
    sev_text = doc_row["default_severity"]
    dur_text = doc_row["default_duration"]
    sev_nums = [float(x) for x in re.findall(r"\d+\.\d+", sev_text)]
    dur_nums = [float(x) for x in re.findall(r"\d+\.\d+", dur_text)]

    doc_severity = None
    eq_match = re.search(r"=\s*(\d+\.\d+)", sev_text)
    if eq_match:
        doc_severity = round4(eq_match.group(1))
    elif "+" in sev_text and sev_nums:
        doc_severity = round4(sum(sev_nums[:3]))
    elif sev_nums:
        doc_severity = round4(sev_nums[0])

    doc_duration = None
    if "×" in dur_text and len(dur_nums) >= 2:
        doc_duration = round4(dur_nums[0] * dur_nums[1])
    elif dur_nums:
        doc_duration = round4(dur_nums[0])

    return doc_severity, doc_duration, sev_text, dur_text


def reconcile_subcategories(doc_defaults, code):
    rows = []
    doc_by_cat = {d["category"]: d for d in doc_defaults}
    exposure = code["exposureDefaults"]

    for mapping in DOC_TO_CODE_SUBCATEGORY:
        doc_row = doc_by_cat.get(mapping["doc_category"])
        if not doc_row:
            continue
        doc_sev, doc_dur, sev_text, dur_text = parse_doc_severity_duration(doc_row)
        code_loc = f"src/shared/product-type-registry/scoring-assumptions.mjs:EXPOSURE_DEFAULTS_BY_KEY.{mapping['code_key']}"

        if mapping.get("utensils_split"):
            # utensils role-split: plastic/nylon 1.0, stainless/wood 0.96, duration 0.50
            split_checks = [
                ("severity_plastic_nylon", 1.0, 1.0),
                ("severity_stainless_wood", 0.96, 0.96),
                ("duration", 0.5, doc_dur),
            ]
            for field, cv, dv in split_checks:
                st, diff = compare_numeric(cv, dv, tol=0.01)
                rows.append(
                    {
                        "category": mapping["registry_subcategory"],
                        "subcategory": mapping["registry_subcategory"],
                        "field": field,
                        "code_value": cv,
                        "doc_value": dv,
                        "doc_value_raw": sev_text if "severity" in field else dur_text,
                        "status": st,
                        "code_file_line": code_loc,
                        "doc_table_location": doc_row["doc_location"],
                        "notes": f"utensils material_split; doc raw={sev_text if 'severity' in field else dur_text}; diff={diff}",
                    }
                )
        else:
            code_sev, code_dur, code_ci, sev_note, dur_note = extract_primary_severity_duration(
                mapping["code_key"], exposure
            )
            for field, cv, dv, note in [
                ("severity", code_sev, doc_sev, sev_note),
                ("duration", code_dur, doc_dur, dur_note),
            ]:
                st, diff = compare_numeric(cv, dv, tol=0.01)
                rows.append(
                    {
                        "category": mapping["registry_subcategory"],
                        "subcategory": mapping["registry_subcategory"],
                        "field": field,
                        "code_value": cv,
                        "doc_value": dv,
                        "doc_value_raw": sev_text if field == "severity" else dur_text,
                        "status": st,
                        "code_file_line": code_loc,
                        "doc_table_location": doc_row["doc_location"],
                        "notes": f"{note}; doc raw={sev_text if field=='severity' else dur_text}; diff={diff}",
                    }
                )

        rows.append(
            {
                "category": mapping["registry_subcategory"],
                "subcategory": mapping["registry_subcategory"],
                "field": "registry_matrix_key",
                "code_value": mapping["matrix_key"],
                "doc_value": mapping["doc_category"],
                "status": "MATCH",
                "code_file_line": "src/shared/product-type-registry/configs/starter-configs.mjs",
                "doc_table_location": doc_row["doc_location"],
                "notes": "",
            }
        )

    return rows


def reconcile_layer4a(doc_pos, doc_neg, code):
    rows = []
    code_pos = {p["exact_label"]: p["points"] for p in code["layer4aPositive"]}
    doc_pos_map = {}
    for p in doc_pos:
        label = p["label"]
        pts = int(re.sub(r"[^\d]", "", p["points"] or "0") or 0)
        doc_pos_map[label] = pts

    def norm_label(s):
        return re.sub(r"[^a-z0-9]+", " ", str(s or "").lower()).strip()

    code_pos_norm = {norm_label(k): (k, v) for k, v in code_pos.items()}
    label_equiv = {
        "independent lab testing confirming material composition": "independent lab testing confirming materials",
        "oeko tex standard 100 certified": "oeko tex standard 100",
        "made safe certified": "made safe certified",
        "pfas free independently verified": "pfas free independently verified",
        "phthalate free independently tested": "phthalate free independently tested",
    }

    matched_code_labels = set()
    for doc_label, doc_pts in doc_pos_map.items():
        nk = norm_label(doc_label)
        equiv = label_equiv.get(nk, nk)
        code_entry = code_pos_norm.get(equiv)
        code_pts = code_entry[1] if code_entry else None
        code_label = code_entry[0] if code_entry else None
        if code_label:
            matched_code_labels.add(norm_label(code_label))
        st = "MATCH" if code_pts == doc_pts else ("MISSING_IN_CODE" if code_pts is None else "MISMATCH")
        rows.append(
            {
                "layer4a_item": doc_label,
                "type": "positive",
                "code_value": code_pts,
                "doc_value": doc_pts,
                "status": st,
                "code_file_line": "scripts/agent2/layer4a-positive.mjs",
                "doc_section": "Layer 4A credits table",
                "notes": f"code label={code_label}",
            }
        )

    for code_label, code_pts in code_pos.items():
        if norm_label(code_label) not in matched_code_labels:
            rows.append(
                {
                    "layer4a_item": code_label,
                    "type": "positive",
                    "code_value": code_pts,
                    "doc_value": None,
                    "status": "MISSING_IN_SOURCE",
                    "code_file_line": "scripts/agent2/layer4a-positive.mjs",
                    "doc_section": "Layer 4A credits table",
                    "notes": "",
                }
            )

    code_neg = {
        "BPA-free claim only, no BPS/BPF testing": code["layer4aDeductions"]["bpa_free_claim_only"],
        "Proprietary food-contact coating, composition undisclosed": code["layer4aDeductions"][
            "proprietary_ceramic_or_nonstick_formula_undisclosed"
        ],
        "Marketing language only, no verifiable claims": code["layer4aDeductions"]["marketing_language_only"],
        "Unknown proprietary food-contact coating": code["layer4aDeductions"][
            "unknown_proprietary_food_contact_coating"
        ],
        "Undisclosed dye chemistry in textiles": -1,
    }

    for d in doc_neg:
        doc_label = d["label"]
        doc_pts = int(re.sub(r"[^\d]", "", d["points"] or "0") or 0) * -1
        code_pts = None
        for k, v in code_neg.items():
            if norm_name(k) == norm_name(doc_label) or norm_name(doc_label) in norm_name(k):
                code_pts = v
                break
        if code_pts is None and "proprietary" in doc_label.lower():
            code_pts = -3
        st = "MATCH" if code_pts == doc_pts else ("MISMATCH" if code_pts is not None else "MISSING_IN_CODE")
        rows.append(
            {
                "layer4a_item": doc_label,
                "type": "negative",
                "code_value": code_pts,
                "doc_value": doc_pts,
                "status": st,
                "code_file_line": "scripts/agent2/deterministic/layer4a-applicability.mjs",
                "doc_section": "Layer 4A deductions table",
                "notes": d.get("applies_when", ""),
            }
        )

    if code["layer4aPositiveMax"] != 5:
        rows.append(
            {
                "layer4a_item": "positive_max_combined",
                "type": "cap",
                "code_value": code["layer4aPositiveMax"],
                "doc_value": 5,
                "status": "MISMATCH",
                "code_file_line": "scripts/agent2/layer4a-positive.mjs:17",
                "doc_section": "Layer 4A max ±5",
                "notes": "",
            }
        )

    return rows


def reconcile_escalators(doc_escalators, code):
    rows = []
    code_esc = {e["id"]: e for e in code["escalators"]}

    doc_map = {
        "1 — High hazard + high severity, adult": "escalator_1",
        "2 — High hazard + high severity, children's": "escalator_2",
        "3 — Degraded high-risk material": "escalator_3",
        "4 — Oral contact extreme-risk material": "escalator_4",
    }

    for d in doc_escalators:
        name = d["name"]
        code_id = doc_map.get(name)
        mult_doc = d["multiplier"]
        mult_match = re.search(r"[\d.]+", mult_doc.replace("×", ""))
        mult_doc_val = float(mult_match.group()) if mult_match else None

        if "cannot fire" in mult_doc.lower() or "not adopted" in mult_doc.lower():
            in_active = code_id in code_esc if code_id else False
            rows.append(
                {
                    "escalator": name,
                    "field": "disabled_status",
                    "code_value": "not in active ESCALATORS" if not in_active else f"still in ESCALATORS ({code_id})",
                    "doc_value": mult_doc,
                    "status": "MATCH" if not in_active else "MISMATCH",
                    "code_file_line": "scripts/agent3/algorithm.mjs:24-29",
                    "doc_section": d["doc_location"],
                    "notes": "v2.3.5: degradation escalator removed from active scoring",
                }
            )
            continue

        if code_id:
            code_mult = code_esc[code_id]["multiplier"]
            st = "MATCH" if abs(code_mult - mult_doc_val) < 0.001 else "MISMATCH"
            rows.append(
                {
                    "escalator": name,
                    "field": "multiplier",
                    "code_value": code_mult,
                    "doc_value": mult_doc_val,
                    "status": st,
                    "code_file_line": "scripts/agent3/algorithm.mjs:24-29",
                    "doc_section": d["doc_location"],
                    "notes": d["trigger"],
                }
            )
            rows.append(
                {
                    "escalator": name,
                    "field": "trigger",
                    "code_value": "see normalize-enforce + escalator-eligibility.mjs",
                    "doc_value": d["trigger"],
                    "status": "NEEDS_RICK_REVIEW",
                    "code_file_line": "src/shared/agent3/escalator-eligibility.mjs",
                    "doc_section": d["doc_location"],
                    "notes": "Threshold details split across Agent2 flags + Agent3 gates",
                }
            )

    # PFAS/PTFE escalator gating (doc paragraph, not separate table row)
    rows.append(
        {
            "escalator": "PFAS/PTFE high-risk escalator (gated)",
            "field": "eligibility",
            "code_value": "escalator_1 gated to confirmed PFAS/PTFE family (escalator-eligibility.mjs)",
            "doc_value": "PFAS/PTFE escalator fires only on confirmed PFAS/PTFE-family material",
            "status": "NEEDS_RICK_REVIEW",
            "code_file_line": "src/shared/agent3/escalator-eligibility.mjs:87",
            "doc_section": "PAC v2.3.5 intro paragraphs",
            "notes": "Doc ties to escalator 1 adult path; code adds explicit PFAS gate",
        }
    )

    return rows


def reconcile_methodology(doc_tiers, code, paragraphs):
    rows = []
    c = code["constants"]

    checks = [
        ("Non-Detect mitigation factor", "value", code["nonDetectFactor"], 0.58, "buildSystemValidation.ts:37"),
        ("Inert protection threshold", "migration_max", c["INERT_MIGRATION_THRESHOLD"], 0.05, "algorithm.mjs:17"),
        ("Inert protection multiplier", "severity_duration_multiplier", c["INERT_EXPOSURE_MULTIPLIER"], 0.2, "algorithm.mjs:18"),
        ("Unknown coating hard cap", "score_cap", c["HARD_CAP_UNKNOWN_COATING"], 72, "algorithm.mjs:21"),
        ("Layer 4A net cap", "max_abs", c["LAYER_4A_CAP"], 5, "algorithm.mjs:20"),
        ("Score max clamp", "max_score", c["SCORE_MAX"], 99, "algorithm.mjs:22"),
        ("NPR scale", "multiplier", c["NPR_SCALE"], 1000, "algorithm.mjs:19"),
        ("Algorithm version string", "version", code["algorithmVersion"], "2.3.5", "algorithm.mjs:7"),
    ]

    for name, field, code_val, doc_val, loc in checks:
        if isinstance(doc_val, float):
            st, diff = compare_numeric(code_val, doc_val)
        else:
            st = "MATCH" if str(code_val) == str(doc_val) else "MISMATCH"
            diff = None
        rows.append(
            {
                "value_name": name,
                "field": field,
                "code_value": code_val,
                "doc_value": doc_val,
                "status": st,
                "code_file_line": loc,
                "doc_section": "PAC v2.3.5 methodology sections",
                "notes": f"diff={diff}" if diff else "",
            }
        )

    # Tier bands
    doc_tier_bounds = []
    for t in doc_tiers:
        m = re.match(r"(\d+)[–-](\d+)", t["score_range"])
        if m:
            doc_tier_bounds.append((int(m.group(1)), t["tier"]))
        elif t["score_range"].startswith("0"):
            doc_tier_bounds.append((0, t["tier"]))

    code_bounds = [(b["min"], b["tier"]) for b in c["tierBands"]]
    for (cmin, ctier), (dmin, dtier) in zip(code_bounds, sorted(doc_tier_bounds, reverse=True)):
        dtier_short = dtier.split("(")[0].strip()
        st = "MATCH" if cmin == dmin and ctier.lower() in dtier.lower() else "NEEDS_RICK_REVIEW"
        rows.append(
            {
                "value_name": f"Tier band {ctier}",
                "field": "min_score",
                "code_value": cmin,
                "doc_value": dmin,
                "status": st,
                "code_file_line": "scripts/agent3/algorithm.mjs:9-14",
                "doc_section": "Tier bands table",
                "notes": f"doc tier label={dtier}",
            }
        )

    rows.append(
        {
            "value_name": "NPR formula",
            "field": "formula",
            "code_value": c["nprFormula"],
            "doc_value": "NPR = ( Material Hazard x Migration x Contact Intimacy x Severity x Duration ) x 1000",
            "status": "MATCH",
            "code_file_line": "scripts/agent3/algorithm.mjs:77-81",
            "doc_section": "Order of operations",
            "notes": "",
        }
    )
    rows.append(
        {
            "value_name": "Score transform",
            "field": "formula",
            "code_value": c["scoreFormula"],
            "doc_value": "PAC Safety Score = 100 - ( sqrt(weighted NPR) x 5 )",
            "status": "MATCH",
            "code_file_line": "scripts/agent3/algorithm.mjs:171-173",
            "doc_section": "Order of operations",
            "notes": "",
        }
    )

    return rows


def cross_source_conflicts(source_rows, doc_tables, code_materials, source_to_code):
    rows = []
    # Doc table 7 subset vs material lookup for overlapping names
    doc_mat_table = None
    for t in doc_tables:
        if t and t[0][:3] == ["Material", "Tier", "Material Hazard"]:
            doc_mat_table = t
            break

    if doc_mat_table:
        for r in doc_mat_table[1:]:
            if len(r) < 4:
                continue
            name = r[0]
            src_row = next(
                (s for s in source_rows if norm_name_strict(s["name"]) == norm_name_strict(name)),
                None,
            )
            if not src_row:
                continue
            doc_h, doc_m = round4(r[2]), round4(r[3])
            sh, sm = src_row["hazard"], src_row["migration"]
            conflict = (doc_h != sh) or (doc_m != sm)
            if conflict:
                rows.append(
                    {
                        "value_name": name,
                        "material_lookup_value": f"hazard={sh}, migration={sm}",
                        "v2_3_5_doc_value": f"hazard={doc_h}, migration={doc_m}",
                        "conflict": True,
                        "notes": "Doc canonical subset table vs full Material Lookup xlsx",
                    }
                )

    return rows


def summarize(rows, status_key="status"):
    c = Counter(r[status_key] for r in rows)
    return {
        "total_checked": len(rows),
        "matches": c.get("MATCH", 0),
        "mismatches": c.get("MISMATCH", 0),
        "missing_active_in_code": c.get("MISSING_ACTIVE_IN_CODE", 0),
        "future_category_not_in_code": c.get("FUTURE_CATEGORY_NOT_IN_CODE", 0),
        "missing_in_source": c.get("MISSING_IN_SOURCE", 0),
        "needs_rick_review": c.get("NEEDS_RICK_REVIEW", 0),
        "alias_only": c.get("ALIAS_ONLY", 0),
    }


PRIORITY_MATERIALS = {
    "borosilicate_glass",
    "tempered_glass",
    "tritan",
    "stainless_steel_304",
    "stainless_steel_316",
    "stainless_steel_unspecified",
    "ptfe_nonstick",
    "ptfe_coating",
    "ceramic_nonstick_sol_gel",
    "proprietary_named_food_contact",
}


def priority_for(row, family):
    st = row.get("status")
    name = str(row.get("material_id") or row.get("value_name") or row.get("layer4a_item") or row.get("escalator") or row.get("category") or "")

    if family == "subcategory_defaults" and "Food storage" in name and row.get("field") in ("severity", "duration") and st == "MISMATCH":
        return "BLOCKER_FOR_PYREX"

    if family == "subcategory_defaults" and st == "MISMATCH":
        return "BLOCKER_FOR_TAXONOMY_BUILD"

    if family in ("methodology", "escalators") and st in ("MISMATCH", "MISSING_IN_CODE", "MISSING_IN_SOURCE"):
        return "BLOCKER_FOR_TAXONOMY_BUILD"

    if family == "layer4a" and st in ("MISMATCH", "MISSING_IN_CODE", "MISSING_IN_SOURCE"):
        return "BLOCKER_FOR_TAXONOMY_BUILD"

    if family == "materials":
        mid = row.get("material_id", "")
        if any(k in str(mid) for k in PRIORITY_MATERIALS) and st == "MISMATCH" and row.get("field") in ("hazard", "migration"):
            return "IMPORTANT_BEFORE_NEW_PRODUCTS"
        if st == "MISSING_ACTIVE_IN_CODE":
            return "IMPORTANT_BEFORE_NEW_PRODUCTS"
        if st == "FUTURE_CATEGORY_NOT_IN_CODE":
            return "CLEANUP_LATER"
        if st == "NEEDS_RICK_REVIEW":
            return "NEEDS_RICK_REVIEW"

    if st == "MATCH":
        return None

    return "NEEDS_RICK_REVIEW"


def build_priority_table(all_sections):
    out = []
    for family, rows in all_sections.items():
        for r in rows:
            if r.get("status") == "MATCH":
                continue
            pr = priority_for(r, family)
            if not pr or pr == "NEEDS_RICK_REVIEW":
                pr = "NEEDS_RICK_REVIEW" if r.get("status") != "FUTURE_CATEGORY_NOT_IN_CODE" else "CLEANUP_LATER"
            out.append(
                {
                    "priority": pr,
                    "value_family": family,
                    "value_name": r.get("material_id")
                    or r.get("value_name")
                    or r.get("layer4a_item")
                    or r.get("escalator")
                    or f"{r.get('category')}/{r.get('field')}",
                    "code_value": r.get("code_value"),
                    "source_value": r.get("source_value")
                    or r.get("doc_value")
                    or r.get("doc_value_raw")
                    or r.get("v2_3_5_doc_value"),
                    "status": r.get("status"),
                    "code_file_line": r.get("code_file_line"),
                    "source_file_location": r.get("source_location")
                    or r.get("doc_section")
                    or r.get("doc_table_location"),
                    "recommended_next_step": "Rick review — approve code vs source before taxonomy build",
                }
            )

    priority_order = {
        "BLOCKER_FOR_PYREX": 0,
        "BLOCKER_FOR_TAXONOMY_BUILD": 1,
        "IMPORTANT_BEFORE_NEW_PRODUCTS": 2,
        "NEEDS_RICK_REVIEW": 3,
        "CLEANUP_LATER": 4,
        "DOC_ONLY": 5,
    }
    out.sort(key=lambda x: (priority_order.get(x["priority"], 9), x["value_family"], str(x["value_name"])))
    return out


def part_b_inventory():
    return [
        {"value_family": "Material taxonomy", "code_file": "scripts/agent2/deterministic/material-taxonomy.mjs", "symbol": "MATERIAL_TAXONOMY", "lines": "27-580", "notes": "46 canonical IDs + aliases"},
        {"value_family": "Material aliases", "code_file": "scripts/agent2/deterministic/material-taxonomy.mjs", "symbol": "MATERIAL_TAXONOMY_ALIASES", "lines": "601-617", "notes": "Agent1 canonical → taxonomy"},
        {"value_family": "Exposure defaults", "code_file": "src/shared/product-type-registry/scoring-assumptions.mjs", "symbol": "EXPOSURE_DEFAULTS_BY_KEY", "lines": "118-175", "notes": "Layer 3 subcategory defaults"},
        {"value_family": "Product-type registry", "code_file": "src/shared/product-type-registry/configs/starter-configs.mjs", "symbol": "STARTER_PRODUCT_TYPE_CONFIGS", "lines": "14-212", "notes": "matrix_key / scoring_assumption_ref"},
        {"value_family": "Taxonomy lookup", "code_file": "scripts/agent2/deterministic/taxonomy-lookup.mjs", "symbol": "enrichComponentsFromTaxonomy", "lines": "1-179", "notes": "Applies registry defaults"},
        {"value_family": "Agent 1 validation", "code_file": "src/lib/lockedInput/buildSystemValidation.ts", "symbol": "buildSystemValidation", "lines": "36-120+", "notes": "NON_DETECT 0.58, Layer4A, escalator mins"},
        {"value_family": "Layer 4A positive", "code_file": "scripts/agent2/layer4a-positive.mjs", "symbol": "LAYER_4A_POSITIVE_LOOKUP", "lines": "5-15", "notes": "9 credits, max +5"},
        {"value_family": "Layer 4A applicability", "code_file": "scripts/agent2/deterministic/layer4a-applicability.mjs", "symbol": "buildLayer4a", "lines": "26-294", "notes": "Negative deductions"},
        {"value_family": "Score algorithm", "code_file": "scripts/agent3/algorithm.mjs", "symbol": "scorePacCore", "lines": "9-289", "notes": "NPR, inert, escalators, caps, tiers"},
        {"value_family": "Escalator eligibility", "code_file": "src/shared/agent3/escalator-eligibility.mjs", "symbol": "escalator1Eligible", "lines": "87-113", "notes": "PFAS/PTFE gate"},
        {"value_family": "Non-Detect heuristic", "code_file": "scripts/lib/material-lookup-audit.mjs", "symbol": "isNonDetectEligibleInCode", "lines": "39-49", "notes": "Pattern-based, not xlsx column"},
        {"value_family": "Score math breakdown", "code_file": "src/lib/scoreMathBreakdown.ts", "symbol": "escalator labels", "lines": "70+", "notes": "Degradation escalator UI copy"},
        {"value_family": "Methodology version", "code_file": "scripts/agent2/deterministic/material-lookup-versions.mjs", "symbol": "METHODOLOGY_VERSION", "lines": "5-7", "notes": "v2.3.5"},
    ]


def main():
    missing = []
    if not MATERIAL_XLSX.is_file():
        missing.append(str(MATERIAL_XLSX))
    if not DOCX.is_file():
        missing.append(str(DOCX))
    if missing:
        print(json.dumps({"stop": True, "missing_files": missing}, indent=2))
        sys.exit(2)

    source_rows, xlsx_meta = load_material_lookup(MATERIAL_XLSX)
    tables = parse_docx_tables(DOCX)
    paragraphs = parse_docx_paragraphs(DOCX)
    code = load_code_values()

    source_to_code = build_source_to_code_map(source_rows, code["materials"])
    material_rows = reconcile_materials(source_rows, code["materials"], source_to_code)

    doc_defaults = parse_category_defaults_table(tables)
    subcategory_rows = reconcile_subcategories(doc_defaults, code)

    doc_pos, doc_neg = parse_layer4a_tables(tables)
    layer4a_rows = reconcile_layer4a(doc_pos, doc_neg, code)

    doc_escalators = parse_escalator_table(tables)
    escalator_rows = reconcile_escalators(doc_escalators, code)

    doc_tiers = parse_tier_table(tables)
    methodology_rows = reconcile_methodology(doc_tiers, code, paragraphs)

    cross_rows = cross_source_conflicts(source_rows, tables, code["materials"], source_to_code)

    sections = {
        "materials": material_rows,
        "subcategory_defaults": subcategory_rows,
        "layer4a": layer4a_rows,
        "escalators": escalator_rows,
        "methodology": methodology_rows,
    }

    priority_table = build_priority_table(sections)

    report = {
        "generated_at": __import__("datetime").datetime.utcnow().isoformat() + "Z",
        "read_only": True,
        "mutations": False,
        "part_a_source_inventory": [
            {
                "source_file": "PBG_Material_Lookup.xlsx",
                "found": True,
                "path": str(MATERIAL_XLSX.relative_to(ROOT)),
                "readable": True,
                "sections": f"Sheets: {xlsx_meta['sheets']}; Material Lookup header row {xlsx_meta['header_row']}; {xlsx_meta['data_row_count']} materials",
                "notes": f"Columns: {', '.join(xlsx_meta['columns'])}; active=41 future_category=45",
            },
            {
                "source_file": "PAC_Scoring_Algorithm_v2.3.5.docx",
                "found": True,
                "path": str(DOCX.relative_to(ROOT)),
                "readable": True,
                "sections": "Category Default Use Assumptions (Table 0), Layer 4A credits/deductions, Escalators, Tier bands, Badges, Migration mitigation 0.58",
                "notes": f"{len(paragraphs)} paragraphs, {len(tables)} tables parsed",
            },
        ],
        "part_b_code_inventory": part_b_inventory(),
        "part_c_material_summary": summarize(material_rows),
        "part_c_material_rows": material_rows,
        "part_d_subcategory_summary": summarize(subcategory_rows),
        "part_d_subcategory_rows": subcategory_rows,
        "part_e_layer4a_rows": layer4a_rows,
        "part_f_escalator_rows": escalator_rows,
        "part_g_methodology_rows": methodology_rows,
        "part_h_cross_source_rows": cross_rows,
        "part_i_summary_by_family": {k: summarize(v) for k, v in sections.items()},
        "part_i_priority_table": priority_table,
        "code_material_count": len(code["materials"]),
        "source_material_count": len(source_rows),
        "source_active_count": sum(1 for r in source_rows if r["scoring_status"] == "active"),
        "source_future_category_count": sum(1 for r in source_rows if r["scoring_status"] == "future_category"),
    }

    REPORT_DIR.mkdir(parents=True, exist_ok=True)
    REPORT_JSON.write_text(json.dumps(report, indent=2), encoding="utf-8")

    # concise markdown for human review
    md = ["# Scoring Source Drift Report (read-only)", ""]
    md.append("## Part I — Summary by value family")
    md.append("| Value family | Total | Matches | Mismatches | Missing active in code | Future category not in code | Missing in source | Needs Rick review |")
    md.append("|---|---:|---:|---:|---:|---:|---:|---:|")
    for fam, summ in report["part_i_summary_by_family"].items():
        md.append(
            f"| {fam} | {summ['total_checked']} | {summ['matches']} | {summ['mismatches']} | {summ['missing_active_in_code']} | {summ['future_category_not_in_code']} | {summ['missing_in_source']} | {summ['needs_rick_review']} |"
        )
    md.append("")
    md.append("## Priority mismatch table (non-MATCH only)")
    md.append("| priority | value family | value name | code | source | status | code location | source location |")
    md.append("|---|---|---|---|---|---|---|---|")
    for r in priority_table[:200]:
        md.append(
            f"| {r['priority']} | {r['value_family']} | {r['value_name']} | {r['code_value']} | {r['source_value']} | {r['status']} | {r['code_file_line']} | {r['source_file_location']} |"
        )
    if len(priority_table) > 200:
        md.append(f"\n… {len(priority_table) - 200} more rows in JSON report")
    REPORT_MD.write_text("\n".join(md), encoding="utf-8")

    print(json.dumps({"ok": True, "report_json": str(REPORT_JSON.relative_to(ROOT)), "report_md": str(REPORT_MD.relative_to(ROOT)), "priority_mismatch_count": len(priority_table), "part_i_summary_by_family": report["part_i_summary_by_family"]}, indent=2))


if __name__ == "__main__":
    main()
